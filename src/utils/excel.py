# excel.py

"""
Пред-обработка Excel-таблиц и вспомогательные операции над ними.
"""
from typing import Optional

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from string_matching import CellClassifier, Match
from utils import Checkpointer


def unique_values_of_whole_sheet(sheet) -> list[str]:
    values = {str(v).strip() for row in sheet.values for v in row if v is not None}  # and v.strip()
    return sorted(values)


def extract_unique_values_to_txt(filename_xlsx_in, filename_txt_out=None):
    wb = load_workbook(filename=filename_xlsx_in, read_only=True)
    sh = wb.active
    values = unique_values_of_whole_sheet(sh)
    filename_out = filename_txt_out or filename_xlsx_in.replace('.xlsx', ' - unique-values.txt')
    with open(filename_out, 'w') as f:
        f.write('\n'.join(values) + '\n')


classifier = None


def classify_cell_content(content: str) -> list[Match]:
    global classifier
    classifier = classifier or CellClassifier()
    matches = classifier.match(content)
    return matches


def colorize_values_on_sheet(sheet) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            content = str(cell.value).strip()
            if content and content != 'None':
                matches = classify_cell_content(content)
                if matches:
                    best = matches[0]
                    class_name, confidence = best.pattern.content_class.description, best.confidence
                    if confidence == 1.0 and 'clear' in best.pattern.content_class.update_content:
                        cell.value = '-'
                    if confidence >= 0.8 and 'replace_with_preprocessed' in best.pattern.content_class.update_content:
                        cell.value = best.cell_text

                    comment = '\n'.join(
                        f'{m.pattern.content_class.description}: {round(m.confidence * 100)}'
                        for m in matches
                        if m.confidence >= confidence * 0.4  # filter too low confidence
                    )
                else:
                    confidence = 0.0
                    comment = 'Не определено'
                # comment = f'{class_name}: {round(confidence * 100)}'
                if confidence <= 0.5:
                    color = 'FF0000'
                elif confidence >= 0.85:
                    color = '00FF00'
                else:
                    color = 'FFCC00'  # thick yellow

                cell.fill = PatternFill("solid", fgColor=color)
                comment = Comment(text=comment, author='Cell classifier')
                # Assign the comment to the cell.
                try:
                    cell.comment = comment
                except AttributeError:
                    # print('failed to add a comment to cell', cell.coordinate)
                    pass


def mark_recognized_values_in_sheet(filename_xlsx_in, filename_txt_out=None):
    wb = load_workbook(filename=filename_xlsx_in)
    sh = wb.active
    # update content
    colorize_values_on_sheet(sh)
    # save back
    filename_out = filename_txt_out or filename_xlsx_in.replace('.xlsx', ' - cells-annotated.xlsx')
    wb.save(filename_out)
    print('Saved file with cells annotated as', filename_out)


def main():
    paths = (
        r'c:\Users\Student\Downloads\ОН_ФТПП_3 курс.xlsx',
    )

    ch = Checkpointer()
    for filename in paths:
        # extract_unique_values_to_txt(filename)
        mark_recognized_values_in_sheet(filename)

        ch.hit('one file processed')

    ch.since_start('all files processed in')


if __name__ == '__main__':
    main()
