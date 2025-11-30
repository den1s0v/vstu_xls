from __future__ import annotations

import json
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Iterable, Mapping, Sequence

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Border, Side

from geom2d import Box
from grammar2d.Match2d import Match2d
from grid import Grid

try:
    from converters.xlsx import ExcelGrid
except ModuleNotFoundError:  # pragma: no cover - optional dependency
    ExcelGrid = None  # type: ignore[assignment]


DEFAULT_COLOR_PALETTE = [
    # "FFF8B4",  # light yellow - excluded to avoid conflict with existing document backgrounds
    "FFCCE5",  # light pink
    "C6E0FF",  # light blue
    "D4F4DD",  # light green
    "FFE0CC",  # light orange
    "E8D1FF",  # light purple
    "FFD6A5",  # peach
    "CFE8FF",  # icy blue
    "F9CCCC",  # rose
    "D1FFD6",  # mint
    "E6E6FA",  # lavender
    "F0E68C",  # khaki (warmer yellow alternative)
]


@dataclass(slots=True)
class WaveDebugExporter:
    """Exports wave results to JSON and annotated Excel copies."""

    output_dir: Path
    enable_json: bool = True
    enable_excel: bool = True
    palette: Sequence[str] = field(default_factory=lambda: tuple(DEFAULT_COLOR_PALETTE))

    def export_wave(
            self,
            wave_index: int,
            grid: Grid | None,
            pattern_names: Sequence[str],
            matches: Iterable[Match2d],
            only_indices=()
    ) -> None:
        """Export wave results in configured formats."""
        if only_indices and wave_index not in only_indices:
            return

        if not (self.enable_json or self.enable_excel):
            return

        self.output_dir.mkdir(parents=True, exist_ok=True)

        matches_unique = list(self._deduplicate_matches(matches))

        if self.enable_json:
            self._export_json(wave_index, pattern_names, matches_unique)

        if self.enable_excel and self._grid_supports_excel(grid):
            self._export_excel(wave_index, grid, matches_unique)

    # region JSON -----------------------------------------------------------------------
    def _export_json(self, wave_index: int, pattern_names: Sequence[str], matches: list[Match2d]) -> None:
        target_path = self.output_dir / f"wave_{wave_index:02d}.json"
        
        # Подсчитываем количество совпадений для каждого паттерна
        pattern_counts: dict[str, int] = {name: 0 for name in pattern_names}
        for match in matches:
            pattern_name = match.pattern.name
            if pattern_name in pattern_counts:
                pattern_counts[pattern_name] += 1
        
        data = {
            "wave_index": wave_index,
            "patterns": pattern_counts,
            "matches": [self._serialize_match(match) for match in matches],
        }

        target_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    @staticmethod
    def _serialize_match(match: Match2d) -> Mapping:
        component_entries = []

        if match.component2match:
            for name, child in match.component2match.items():
                component_entries.append({
                    "name": str(name),
                    "pattern": child.pattern.name,
                    "box": WaveDebugExporter._box_dict(child.box),
                })

        precision = match.precision if match.precision is not None else match.calc_precision()

        return {
            "pattern": match.pattern.name,
            "precision": precision,
            "box": WaveDebugExporter._box_dict(match.box),
            "text_content": match.get_text(),
            "component_count": len(component_entries),
            "components": component_entries,
        }

    @staticmethod
    def _box_dict(box: Box | None) -> Mapping:
        if not box:
            return {}
        return {
            "left": box.left,
            "top": box.top,
            "right": box.right,
            "bottom": box.bottom,
            "width": box.w,
            "height": box.h,
        }

    def export_unused_patterns_report(self, unused_by_pattern: Mapping[str, Sequence[Match2d]]) -> None:
        """Экспортирует отчёт о неиспользованных паттернах в JSON."""
        if not self.enable_json:
            return

        target_path = self.output_dir / "unused_patterns.json"
        report_data = {
            "patterns_analyzed": len(unused_by_pattern),
            "patterns_with_unused": len([p for p, matches in unused_by_pattern.items() if matches]),
            "count_unused_by_pattern": {
                pattern_name: len(matches)  # f"{len(matches)} unused matches"
                for pattern_name, matches in unused_by_pattern.items()
            },
            "unused_by_pattern": {
                pattern_name: [
                    self._serialize_match(match) for match in matches
                ]
                for pattern_name, matches in unused_by_pattern.items()
                if matches
            },
        }

        target_path.write_text(json.dumps(report_data, ensure_ascii=False, indent=2), encoding="utf-8")

    # endregion ------------------------------------------------------------------------

    # region Excel ---------------------------------------------------------------------
    def _export_excel(self, wave_index: int, grid: Grid, matches: list[Match2d]) -> None:
        assert ExcelGrid is not None  # for type-checkers
        excel_grid: ExcelGrid = grid  # type: ignore[assignment]

        workbook_copy = self._clone_workbook(excel_grid)
        worksheet_copy = workbook_copy[excel_grid._worksheet.title]
        self._clear_cell_fills(worksheet_copy)
        pattern_colors = self._resolve_colors(matches)
        border_style = self._make_border()

        # Группируем совпадения по верхней левой ячейке
        matches_by_position = self._group_matches_by_position(matches)

        # Подсвечиваем ячейки, выбирая паттерн с максимальной точностью для каждой ячейки
        self._highlight_cells_by_best_match(worksheet_copy, matches, pattern_colors, border_style)

        # Добавляем комментарии в верхние левые ячейки
        self._add_match_annotations(worksheet_copy, matches_by_position)

        target_path = self.output_dir / f"wave_{wave_index:02d}.xlsx"
        workbook_copy.save(target_path)

    @staticmethod
    def _grid_supports_excel(grid: Grid | None) -> bool:
        if grid is None or ExcelGrid is None:
            return False
        return isinstance(grid, ExcelGrid)

    @staticmethod
    def _clone_workbook(grid: "ExcelGrid"):
        """Make a deep copy of the original workbook to avoid altering source files."""
        workbook = grid._worksheet.parent
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return openpyxl.load_workbook(stream)

    @staticmethod
    def _clear_cell_fills(worksheet) -> None:
        """Clear all cell fill backgrounds to remove existing coloring from source document."""
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.fill and cell.fill.fill_type:
                    cell.fill = PatternFill()

    def _resolve_colors(self, matches: list[Match2d]) -> Mapping[str, str]:
        palette_cycle = list(self.palette) or list(DEFAULT_COLOR_PALETTE)
        color_map: dict[str, str] = {}
        index = 0

        for match in matches:
            pattern_name = match.pattern.name
            if pattern_name in color_map:
                continue
            color_map[pattern_name] = self._normalize_color(palette_cycle[index % len(palette_cycle)])
            index += 1

        return color_map

    @staticmethod
    def _normalize_color(color: str) -> str:
        color = color.upper().lstrip("#")
        if len(color) == 6:
            return "FF" + color
        if len(color) == 8:
            return color
        return "FFFF0000"

    @staticmethod
    def _make_border() -> Border:
        side = Side(style="thin", color="FF000000")
        return Border(left=side, right=side, top=side, bottom=side)

    def _highlight_cells_by_best_match(
        self, 
        worksheet, 
        matches: list[Match2d], 
        pattern_colors: Mapping[str, str], 
        border: Border
    ) -> None:
        """Подсвечивает ячейки, выбирая для каждой ячейки паттерн с максимальной точностью."""
        # Словарь: (row, col) -> (match, precision)
        cell_to_best_match: dict[tuple[int, int], tuple[Match2d, float]] = {}
        
        # Проходим по всем совпадениям и для каждой ячейки выбираем лучшее
        for match in matches:
            if not match.box:
                continue
            
            precision = match.precision if match.precision is not None else match.calc_precision()
            if precision is None:
                precision = 0.0
            
            # Проходим по всем ячейкам, покрытым этим совпадением
            for row in range(match.box.top, match.box.bottom):
                for col in range(match.box.left, match.box.right):
                    cell_key = (row, col)
                    
                    # Если для этой ячейки ещё нет совпадения или текущее лучше
                    if cell_key not in cell_to_best_match:
                        cell_to_best_match[cell_key] = (match, precision)
                    else:
                        _, best_precision = cell_to_best_match[cell_key]
                        if precision > best_precision:
                            cell_to_best_match[cell_key] = (match, precision)
        
        # Подсвечиваем ячейки цветом лучшего паттерна
        for (row, col), (best_match, _) in cell_to_best_match.items():
            fill = PatternFill(
                start_color=pattern_colors[best_match.pattern.name],
                end_color=pattern_colors[best_match.pattern.name],
                fill_type="solid",
            )
            cell = worksheet.cell(row=row + 1, column=col + 1)
            cell.fill = fill
            cell.border = border

    @staticmethod
    def _highlight_box(worksheet, box: Box | None, fill: PatternFill, border: Border) -> None:
        if not box:
            return

        for row in range(box.top, box.bottom):
            for col in range(box.left, box.right):
                cell = worksheet.cell(row=row + 1, column=col + 1)
                cell.fill = fill
                cell.border = border

    @staticmethod
    def _group_matches_by_position(matches: list[Match2d]) -> dict[tuple[int, int], list[Match2d]]:
        """Группирует совпадения по их верхней левой ячейке (x, y)."""
        grouped: dict[tuple[int, int], list[Match2d]] = {}

        for match in matches:
            if not match.box:
                continue
            position = (match.box.left, match.box.top)
            if position not in grouped:
                grouped[position] = []
            grouped[position].append(match)

        return grouped

    @staticmethod
    def _format_match_annotation(matches: list[Match2d]) -> str:
        """Формирует текст аннотации для списка совпадений в одной позиции."""
        if not matches:
            return ""

        # Подготавливаем данные для сортировки: (размер, точность, паттерн, размер текстом)
        match_info = []
        for match in matches:
            if not match.box:
                continue
            size = match.box.w * match.box.h
            precision = match.precision if match.precision is not None else match.calc_precision()
            pattern_name = match.pattern.name
            match_info.append((size, precision, pattern_name, match.box.w, match.box.h))

        # Сортируем: сначала по размеру (больше -> меньше), затем по точности (больше -> меньше)
        match_info.sort(key=lambda x: (-x[0], -x[1]))

        # Формируем строки аннотации
        lines = []
        for size, precision, pattern_name, w, h in match_info:
            precision_str = f"{precision:.2f}" if precision is not None else "N/A"
            lines.append(f"{pattern_name}: {w}x{h}, точность {precision_str}")

        return "\n".join(lines)

    @staticmethod
    def _add_match_annotations(worksheet, matches_by_position: dict[tuple[int, int], list[Match2d]]) -> None:
        """Добавляет комментарии-аннотации в верхнюю левую ячейку каждого совпадения."""
        for (col, row), matches in matches_by_position.items():
            if not matches:
                continue

            annotation_text = WaveDebugExporter._format_match_annotation(matches)
            if not annotation_text:
                continue

            # Excel использует 1-based координаты, поэтому добавляем 1
            excel_row = row + 1
            excel_col = col + 1

            try:
                cell = worksheet.cell(row=excel_row, column=excel_col)
                comment = Comment(text=annotation_text, author="Grammar Matcher")
                cell.comment = comment
            except (AttributeError, Exception):
                # Игнорируем ошибки при добавлении комментария
                pass

    # endregion ------------------------------------------------------------------------

    # region Helpers -------------------------------------------------------------------
    @staticmethod
    def _deduplicate_matches(matches: Iterable[Match2d]) -> Iterable[Match2d]:
        seen = set()
        for match in matches:
            marker = id(match)
            if marker in seen:
                continue
            seen.add(marker)
            yield match

    # endregion -----------------------------------------------------------------------

