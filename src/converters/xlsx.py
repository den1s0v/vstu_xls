from pathlib import Path
from typing import Optional, Self

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from converters.abstract import AbstractGridBuilder
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.styles import Color
from geom2d.point import Point
from geom2d.size import Size
from grid import Cell, CellStyle, Grid
from utils.openpyxl_colorconvert import theme_and_tint_to_rgb, get_theme_colors


COORD_PAD = 1  # Use to turn Excel's 1-based coordinates to 0-based indices.
# COORD_PAD = 0  # Use to keep coordinates as-is.


def get_rgb(color: Color, wb: Workbook) -> Optional[str]:
    """
    Получить RGB цвет из объекта Color openpyxl.

    Параметры:
        color (openpyxl.styles.Color): объект цвета (может быть rgb, theme или indexed)
        wb (openpyxl.Workbook): рабочая книга для доступа к темам

    Возвращает:
        str: RGB в шестнадцатеричном формате, например 'FF0000' (без #)
    """
    if color is None:
        return None

    # 1. Если явно задан RGB
    if color.type == "rgb" and color.rgb:
        return color.rgb.upper()
    # 2. Если задан индекс
    elif color.type == "indexed" and color.indexed is not None:
        indexed_rgb = COLOR_INDEX[color.indexed]
        return indexed_rgb.upper()
    # 3. Если задана тема
    elif color.type == "theme" and color.theme is not None:
        return theme_and_tint_to_rgb(
            wb,
            color.theme,
            tint=color.tint,
        ).upper()

    return None  # Если ничего не найдено


class ExcelGrid(Grid, AbstractGridBuilder):
    """A Grid implementation that builds a model from an Excel worksheet using openpyxl.
    Stores references to original openpyxl cells for style and formatting access."""

    def __init__(self, worksheet: Worksheet) -> None:
        super().__init__()
        self._worksheet = worksheet
        self._load_cells(worksheet)

    @classmethod
    def read_xlsx(cls, filepath: str | Path, _sheet=None) -> Self:
        """ Read Grid from the first active sheet from workbook at given path.
         TODO: implement sheet choosing.
         """
        return ExcelGrid(openpyxl.load_workbook(Path(filepath)).active)

    def _load_cells(self, data: Worksheet) -> None:
        """Load cells from the provided worksheet, creating Cell objects and storing openpyxl cell references."""
        # Iterate through all cells in the worksheet's defined dimensions
        for row in data.iter_rows(
            min_row=1,
            max_row=data.max_row,
            min_col=1,
            max_col=data.max_column,
        ):
            for excel_cell in row:
                if excel_cell.value is None:
                    continue  # Skip empty cells

                is_merged = any(
                    excel_cell.coordinate in merged_range
                    for merged_range in data.merged_cells.ranges
                )
                if is_merged:
                    continue

                x = excel_cell.column - COORD_PAD
                y = excel_cell.row - COORD_PAD

                # Create CellStyle from openpyxl cell properties
                style = self._create_cell_style(excel_cell)

                # Create Cell with content and style
                cell = Cell(
                    grid=self,
                    point=Point(x, y),
                    size=Size(1, 1),  # Default size; merged cells are handled below
                    content=str(excel_cell.value),
                    style=style,
                )

                # Store reference to original openpyxl cell in cell.data
                cell.data = {"openpyxl_cell": excel_cell}

                # Register the cell
                self.register_cell(cell)

        # Handle merged cells
        self._process_merged_cells()

    def _create_cell_style(self, excel_cell) -> CellStyle:
        """Create a CellStyle object from openpyxl cell properties."""

        # Initialize style attributes
        font_style = set()
        borders = set()
        background_color = None
        font_color = None

        # Font styles
        if excel_cell.font:
            if excel_cell.font.b:
                font_style.add("bold")
            if excel_cell.font.i:
                font_style.add("italic")
            if excel_cell.font.u:
                font_style.add("underline")

        # Border styles
        if excel_cell.border:
            if excel_cell.border.left and excel_cell.border.left.style:
                borders.add("left")
            if excel_cell.border.right and excel_cell.border.right.style:
                borders.add("right")
            if excel_cell.border.top and excel_cell.border.top.style:
                borders.add("top")
            if excel_cell.border.bottom and excel_cell.border.bottom.style:
                borders.add("bottom")

        # Background color (convert RGB to hex if present)
        if excel_cell.fill and excel_cell.fill.fgColor:
            background_color = get_rgb(excel_cell.fill.fgColor, self._worksheet.parent)

        # Foreground color (font color)
        ### print(excel_cell.font.color)
        if excel_cell.font and excel_cell.font.color:
            font_color = get_rgb(excel_cell.font.color, self._worksheet.parent)

        # Create CellStyle object
        return CellStyle(
            font_style=font_style,
            background_color=background_color,
            borders=borders,
            font_color=font_color,
        )

    def _process_merged_cells(self) -> None:
        """Process merged cell ranges, updating Cell objects with appropriate sizes."""
        if not self._worksheet.merged_cells:
            return

        for merged_range in self._worksheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = (
                merged_range.min_col,
                merged_range.min_row,
                merged_range.max_col,
                merged_range.max_row,
            )
            x = min_col - COORD_PAD
            y = min_row - COORD_PAD
            x2 = max_col - COORD_PAD
            y2 = max_row - COORD_PAD

            point = Point(x, y)

            # Calculate size of the merged cell
            width = x2 - x + 1
            height = y2 - y + 1
            size = Size(width, height)

            # Get the top-left cell from point2cell
            cell = self.get_cell(point)
            if not cell:
                # Create a new cell if none exists (e.g., empty merged cell)
                excel_cell = self._worksheet.cell(row=min_row, column=min_col)
                style = self._create_cell_style(excel_cell)
                cell = Cell(
                    grid=self,
                    point=point,
                    size=size,
                    content=str(excel_cell.value or ""),
                    style=style,
                )
                cell.data = {"openpyxl_cell": excel_cell}
            else:
                # Update existing cell's size
                cell.size = size

            # Re-register the cell to update point2cell mappings
            self.register_cell(cell)

    def supports_cell_merging(self) -> bool:
        """ExcelGrid supports merged cells."""
        return True
