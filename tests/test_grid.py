import unittest

import openpyxl
from tests_bootstrapper import init_testing_environment

init_testing_environment()

from converters.text import TxtGrid
from converters.xlsx import ExcelGrid
from geom2d.box import Box
from geom2d.direction import LEFT, RIGHT, UP, Direction
from utils.path import current_rootpath

testdata_path = current_rootpath() / "tests" / "test_data"


class GridTestCase(unittest.TestCase):
    def test_in_1(self):
        gs = [
            TxtGrid((testdata_path / "grid1.tsv").read_text()),
            ExcelGrid(openpyxl.load_workbook(testdata_path / "grid1.xlsx").active),
        ]
        for g in gs:
            gw = g.get_view()
            print("grid.get_bounding_box:", g.get_bounding_box())
            print("grid-view range:", gw)

            self.assertEqual(tuple(gw), (0, 0, 9, 9))
            self.assertEqual((0, 0, 9, 9), tuple(gw))

            # Справа НАЛЕВО, снизу ВВЕРХ:
            for cell in gw.iterate_cells((LEFT, UP)):
                print(cell)

            col = gw.get_region(Box(0, 0, 1, 9))
            print("col-view range:", col)
            content_list = []
            # for point in col.iterate_points( [UP, RIGHT] ):
            #     # print(point)
            #     ...
            # Снизу ВВЕРХ, слева НАПРАВО:
            for cell in col.iterate_cells([UP, RIGHT]):
                print(cell)
                content_list.append(cell.cell.content)

            col_chars = "".join(content_list)
            print(col_chars)

            # self.assertEqual(col_chars, '87654321')
            self.assertEqual(col_chars, "12345678")

            row = gw.get_region(Box(0, 0, 9, 1))
            print("row-view range:", row)
            content_list = []
            for cell in row.iterate_cells():
                print(cell)
                content_list.append(cell.cell.content)

            row_chars = "".join(content_list)
            print(row_chars)

            self.assertEqual(row_chars, "ABCDEFGH")

            rect = gw.get_region(Box(3, 3, 3, 3))
            print("rect-view range:", rect)
            for d in Direction.known_instances():
                outer = rect.look_outside(d)
                chars = (cell.cell.content for cell in outer.iterate_cells())
                print("rect-view range, %s:" % d.prop_name, outer)
                print("                   ", "".join(chars))

            print()
            rect = gw.get_region(Box(2, 2, 5, 5))
            print("rect-view range:", rect)
            for d in Direction.known_instances():
                outer = rect.look_outside(d, distance=2)
                chars = (cell.cell.content for cell in outer.iterate_cells())
                print("rect-view range, %s:" % d.prop_name, outer)
                print("                   ", "".join(chars))

            print()
            print("# zero distance")
            rect = gw.get_region(Box(2, 2, 4, 4))
            print("rect-view range:", rect)
            for d in Direction.known_instances():
                outer = rect.look_outside(d, distance=0)
                # chars = (cell.cell.content for cell in outer.iterateCells())
                print("rect-view range, %s:" % d.prop_name, outer)
                # print('                   ', ''.join(chars))

            print()
            print("# ignored distance")
            rect = gw.get_region(Box(2, 2, 4, 4))
            print("rect-view range:", rect)
            for d in Direction.known_instances():
                outer = rect.look_outside(d, distance=-10)  # ignored distance
                chars = (cell.cell.content for cell in outer.iterate_cells())
                print("rect-view range, %s:" % d.prop_name, outer)
                print("                   ", "".join(chars))

            print()
            print("# too large distance, clamp.")
            rect = gw.get_region(Box(2, 2, 4, 4))
            print("rect-view range:", rect)
            for d in Direction.known_instances():
                outer = rect.look_outside(d, distance=999)  # too large distance, clamp.
                chars = (cell.cell.content for cell in outer.iterate_cells())
                print("rect-view range, %s:" % d.prop_name, outer)
                print("                   ", "".join(chars))

    def test_in_2(self):
        workbook = openpyxl.load_workbook(testdata_path / "grid2.xlsx")
        g = ExcelGrid(workbook.active)
        gw = g.get_view()
        self.assertEqual(gw.get_cell_view((0, 0)).cell.content, "a")
        self.assertEqual(gw.get_cell_view((0, 1)), None)
        self.assertEqual(gw.get_cell_view((0, 2)).cell.content, "x")
        self.assertEqual(gw.get_cell_view((1, 2)).cell.content, "m")
        self.assertEqual(gw.get_cell_view((2, 0)).cell.content, "b")
        self.assertEqual(gw.get_cell_view((2, 1)).cell.content, "y")

        self.assertEqual(gw.get_cell_view((3, 0)).cell.content, "merged")
        self.assertEqual(gw.get_cell_view((3, 0)).size.h, 3)
        self.assertEqual(gw.get_cell_view((3, 0)).size.w, 2)

        self.assertIs(gw.get_cell_view((3, 0)), gw.get_cell_view((3, 1)))
        self.assertIs(gw.get_cell_view((3, 0)), gw.get_cell_view((3, 2)))
        self.assertIs(gw.get_cell_view((3, 0)), gw.get_cell_view((4, 0)))
        self.assertIs(gw.get_cell_view((3, 0)), gw.get_cell_view((4, 1)))
        self.assertIs(gw.get_cell_view((3, 0)), gw.get_cell_view((4, 2)))

    def test_in_3(self):
        workbook = openpyxl.load_workbook(testdata_path / "grid3.xlsx")
        g = ExcelGrid(workbook.active)
        gw = g.get_view()
        self.assertEqual(gw.get_cell_view((0, 0)).cell.content, "a")
        self.assertEqual(gw.get_cell_view((1, 0)).cell.content, "b")
        self.assertEqual(gw.get_cell_view((2, 0)).cell.content, "c")
        self.assertEqual(gw.get_cell_view((3, 0)).cell.content, "d")
        self.assertEqual(gw.get_cell_view((4, 0)).cell.content, "e")
        self.assertNotEqual(gw.get_cell_view((4, 0)).cell.style.font_color, None)

        self.assertEqual(gw.get_cell_view((5, 0)), None)

        self.assertEqual(gw.get_cell_view((0, 1)).cell.content, "x")
        self.assertEqual(gw.get_cell_view((1, 1)).cell.content, "y")
        self.assertNotEqual(gw.get_cell_view((1, 1)).cell.style.font_color, None)
        self.assertEqual(gw.get_cell_view((2, 1)).cell.content, "z")
        self.assertIn("right", gw.get_cell_view((2, 1)).cell.style.borders)

        self.assertEqual(gw.get_cell_view((3, 1)).cell.content, "i")
        self.assertIn("bold", gw.get_cell_view((3, 1)).cell.style.font_style)

        self.assertEqual(gw.get_cell_view((4, 1)).cell.content, "j")
        self.assertEqual(gw.get_cell_view((4, 1)).size.h, 2)
        self.assertEqual(gw.get_cell_view((4, 1)).size.w, 1)
        self.assertEqual(gw.get_cell_view((4, 2)).cell.content, "j")
        self.assertIn("italic", gw.get_cell_view((4, 2)).cell.style.font_style)

        self.assertEqual(gw.get_cell_view((5, 1)).cell.content, "k")

        for i, j in ([0, 0], [0, 1], [1, 0], [1, 1]):
            self.assertEqual(gw.get_cell_view((1 + i, 2 + j)).cell.content, "m", f'i,j={i, j}')
            self.assertEqual(gw.get_cell_view((1 + i, 2 + j)).size.h, 2, f'i,j={i, j}')
            self.assertEqual(gw.get_cell_view((1 + i, 2 + j)).size.w, 2, f'i,j={i, j}')
        for i, j in ([0, 2], [1, 2], [2, 2], [2, 1], [2, 0]):
            self.assertTrue(not (cw := gw.get_cell_view((1 + i, 2 + j))) or cw.cell.content != "m", f'i,j={i, j}')

        self.assertEqual(gw.get_cell_view((3, 2)).cell.content, "s")
        self.assertEqual(gw.get_cell_view((5, 2)).cell.content, "_")

        self.assertEqual(gw.get_cell_view((0, 3)).cell.content, "t")
        self.assertNotEqual(gw.get_cell_view((0, 3)).cell.style.background_color, "00000000")

        self.assertEqual(gw.get_cell_view((3, 3)).cell.content, "w")
        self.assertEqual(gw.get_cell_view((4, 3)).cell.content, "p")
        self.assertIn("bold", gw.get_cell_view((4, 3)).cell.style.font_style)
        self.assertIn("italic", gw.get_cell_view((4, 3)).cell.style.font_style)
        self.assertIsNone(gw.get_cell_view((5, 3)))


if __name__ == "__main__":
    # unittest.main()
    GridTestCase.test_in_2(...)
